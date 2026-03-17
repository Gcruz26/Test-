import csv
from datetime import date, datetime, time
from io import BytesIO, StringIO
from typing import Any

import xlrd
from openpyxl import load_workbook


def _normalize_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def _build_headers(raw_headers: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}

    for idx, header in enumerate(raw_headers, start=1):
        base = str(header).strip() if header is not None and str(header).strip() else f"column_{idx}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        headers.append(base if count == 0 else f"{base}_{count + 1}")

    return headers


def _rows_to_payloads(rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []

    headers = _build_headers(rows[0])
    payloads: list[dict[str, Any]] = []

    for row in rows[1:]:
        normalized = [_normalize_value(value) for value in row]
        if len(normalized) < len(headers):
            normalized.extend([None] * (len(headers) - len(normalized)))
        elif len(normalized) > len(headers):
            extra_headers = [f"extra_column_{i}" for i in range(len(headers) + 1, len(normalized) + 1)]
            headers = headers + extra_headers

        payload = {headers[i]: normalized[i] for i in range(len(headers))}
        if all(value in (None, "") for value in payload.values()):
            continue
        payloads.append(payload)

    return payloads


def parse_csv(content: bytes) -> list[dict[str, Any]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.reader(StringIO(text))
    rows = [row for row in reader]
    return _rows_to_payloads(rows)


def parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
    worksheet = workbook.active
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    return _rows_to_payloads(rows)


def parse_xls(content: bytes) -> list[dict[str, Any]]:
    workbook = xlrd.open_workbook(file_contents=content)
    sheet = workbook.sheet_by_index(0)
    rows: list[list[Any]] = []

    for row_idx in range(sheet.nrows):
        row_values: list[Any] = []
        for col_idx in range(sheet.ncols):
            value = sheet.cell_value(row_idx, col_idx)
            row_values.append(value)
        rows.append(row_values)

    return _rows_to_payloads(rows)


def parse_report(content: bytes, file_format: str) -> list[dict[str, Any]]:
    fmt = file_format.lower()
    if fmt == "csv":
        return parse_csv(content)
    if fmt in {"xlsx", "xlsm"}:
        return parse_xlsx(content)
    if fmt == "xls":
        return parse_xls(content)
    raise ValueError(f"Unsupported file format: {file_format}")
